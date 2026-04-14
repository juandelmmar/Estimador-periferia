"""
Backend - Estimador de Proyectos con IA (Edición Producción)
==========================================================
Cliente: Periferia IT Group
Tecnologías: Flask + Groq (Llama 3.3) + OpenPyXL
"""

import os
import json
import tempfile
import openpyxl
from flask import Flask, request, jsonify, send_file, send_from_directory
from groq import Groq
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

app = Flask(__name__, static_folder=".", static_url_path="")

# ============ CONFIGURACIÓN DE SEGURIDAD ============
# En local: Puedes setear esto en tu terminal con $env:GROQ_API_KEY="tu_llave"
# En Render: Agrégala en la pestaña 'Environment'
api_key = os.environ.get("GROQ_API_KEY")
client = Groq(api_key=api_key)
MODELO = "llama-3.3-70b-versatile"

PROMPT_SISTEMA = """Actúa como un líder técnico senior con experiencia en estimación de proyectos de software.
Tu objetivo es generar una estimación de horas justa y realista.

REGLAS CRÍTICAS:
1. Las horas deben ser JUSTAS: ni infladas ni demasiado bajas.
2. El rango típico por actividad es entre 4 y 40 horas.
3. Responde SOLO en formato JSON exacto, sin markdown ni texto extra.

ESTRUCTURA JSON:
{
  "cliente": "Nombre",
  "ingeniero": "Nombre del ingeniero",
  "backend": "Tecnologías",
  "frontend": "Tecnologías",
  "base_datos": "DB",
  "cloud": "Proveedor Cloud (AWS, Azure, GCP, etc.)",
  "actividades": [
    {"actividad": "Nombre", "descripcion": "Detalle", "funcionalidades": "Funciones", "horas": 8}
  ],
  "pruebas_pct": 15,
  "entendimiento_pct": 10,
  "riesgo_pct": 5,
  "notas": ["nota1"]
}"""

@app.route("/")
def index():
    return send_from_directory(".", "index.html")

@app.route("/api/estimar", methods=["POST"])
def estimar():
    data = request.get_json()
    descripcion = data.get("descripcion", "")
    cliente = data.get("cliente", "")
    ingeniero = data.get("ingeniero", "")

    if not descripcion or len(descripcion) < 20:
        return jsonify({"error": "La descripción es muy corta"}), 400

    if not api_key:
        return jsonify({"error": "API Key no configurada en el servidor"}), 500

    try:
        parts = []
        if cliente: parts.append(f"Cliente: {cliente}")
        if ingeniero: parts.append(f"Ingeniero a cargo: {ingeniero}")
        parts.append(f"Descripción: {descripcion}")
        prompt_usuario = "\n".join(parts)

        # Uso del cliente oficial de Groq
        completion = client.chat.completions.create(
            model=MODELO,
            messages=[
                {"role": "system", "content": PROMPT_SISTEMA},
                {"role": "user", "content": prompt_usuario}
            ],
            temperature=0.2,
            max_tokens=2048
        )

        respuesta_ia = completion.choices[0].message.content
        # Limpieza por si la IA devuelve markdown
        clean_json = respuesta_ia.replace("```json", "").replace("```", "").strip()
        parsed = json.loads(clean_json)
        # Override with user-provided values
        if cliente: parsed['cliente'] = cliente
        if ingeniero: parsed['ingeniero'] = ingeniero
        
        return jsonify(parsed)

    except Exception as e:
        return jsonify({"error": f"Error al procesar con IA: {str(e)}"}), 500

@app.route("/api/descargar-excel", methods=["POST"])
def descargar_excel():
    data = request.get_json()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estimación"

    # ── Color palette (Periferia brand) ──────────────────────────────
    GREEN_DARK   = "15601D"
    GREEN_MED    = "1E7A28"
    GREEN_NEON   = "6DFD8C"
    GREEN_LIGHT  = "CCFFD6"
    GREEN_PALE   = "F0FFF3"
    DARK         = "212121"
    WHITE        = "FFFFFF"
    GRAY_LIGHT   = "F5F5F5"
    GRAY_MID     = "E0E0E0"
    ACCENT_TEAL  = "1B5E20"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(bold=False, color="212121", size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

    def border_thin(sides="all"):
        thin = Side(style="thin", color="CCCCCC")
        none = Side(style=None)
        if sides == "all":
            return Border(left=thin, right=thin, top=thin, bottom=thin)
        if sides == "bottom":
            return Border(bottom=thin)
        if sides == "outer":
            thick = Side(style="medium", color=GREEN_DARK)
            return Border(left=thick, right=thick, top=thick, bottom=thick)
        return Border()

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ── Column widths ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 12

    # ── Row height helpers ────────────────────────────────────────────
    def row_h(row, h):
        ws.row_dimensions[row].height = h

    # ═══════════════════════════════════════════
    # HEADER BANNER  (rows 1-5)
    # ═══════════════════════════════════════════
    row_h(1, 8)
    row_h(2, 48)
    row_h(3, 22)
    row_h(4, 22)
    row_h(5, 10)

    # Banner background A1:E5
    for r in range(1, 6):
        for c in range(1, 6):
            cell = ws.cell(r, c)
            cell.fill = fill(GREEN_DARK)

    # Company name
    ws.merge_cells("B2:D2")
    c = ws["B2"]
    c.value = "PERIFERIA IT GROUP"
    c.font = Font(bold=True, color=GREEN_NEON, size=22, name="Calibri")
    c.alignment = align("left", "center")
    c.fill = fill(GREEN_DARK)

    ws.merge_cells("E2:E2")
    ws["E2"].fill = fill(GREEN_DARK)

    ws.merge_cells("B3:D3")
    c = ws["B3"]
    c.value = "Estimación Profesional de Proyectos de Software"
    c.font = Font(bold=False, color=GREEN_LIGHT, size=11, name="Calibri", italic=True)
    c.alignment = align("left", "center")
    c.fill = fill(GREEN_DARK)

    from datetime import datetime
    ws.merge_cells("B4:D4")
    c = ws["B4"]
    c.value = f"Generado el {datetime.now().strftime('%d/%m/%Y  %H:%M')}"
    c.font = Font(color="88BB99", size=9, name="Calibri")
    c.alignment = align("left", "center")
    c.fill = fill(GREEN_DARK)

    # ═══════════════════════════════════════════
    # PROJECT INFO BLOCK  (rows 6-13)
    # ═══════════════════════════════════════════
    row_h(6, 8)
    row_h(7, 28)
    row_h(8, 24)
    row_h(9, 24)
    row_h(10, 24)
    row_h(11, 24)
    row_h(12, 24)
    row_h(13, 10)

    # Section header
    ws.merge_cells("B7:E7")
    c = ws["B7"]
    c.value = "  INFORMACIÓN DEL PROYECTO"
    c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    c.fill = fill(GREEN_MED)
    c.alignment = align("left", "center")

    info_fields = [
        ("CLIENTE",    data.get("cliente", "—")),
        ("INGENIERO",  data.get("ingeniero", "—")),
        ("BACKEND",    data.get("backend", "—")),
        ("FRONTEND",   data.get("frontend", "—")),
        ("BASE DATOS", data.get("base_datos", "—")),
        ("CLOUD",      data.get("cloud", "—")),
    ]

    for i, (label, value) in enumerate(info_fields):
        r = 8 + i
        # Label cell
        lc = ws.cell(r, 2, value=f"  {label}")
        lc.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        lc.fill = fill(ACCENT_TEAL)
        lc.alignment = align("left", "center")
        lc.border = Border(bottom=Side(style="thin", color="2E8B40"))
        # Value cell
        ws.merge_cells(f"C{r}:E{r}")
        vc = ws.cell(r, 3, value=f"  {value}")
        vc.font = Font(color=DARK, size=10, name="Calibri")
        vc.fill = fill(GREEN_PALE if i % 2 == 0 else WHITE)
        vc.alignment = align("left", "center")
        vc.border = Border(bottom=Side(style="thin", color=GRAY_MID))

    # ═══════════════════════════════════════════
    # STATS BOXES  (rows 14-18)
    # ═══════════════════════════════════════════
    row_h(14, 8)
    row_h(15, 32)
    row_h(16, 20)
    row_h(17, 20)
    row_h(18, 8)

    actividades = data.get("actividades", [])
    total_dev = sum(a.get("horas", 0) for a in actividades)
    p_pruebas  = data.get("pruebas_pct", 0) / 100
    p_ent      = data.get("entendimiento_pct", 0) / 100
    p_riesgo   = data.get("riesgo_pct", 0) / 100
    total_final = round(total_dev * (1 + p_pruebas + p_ent + p_riesgo), 1)

    # Stat box 1 - Horas Dev
    ws.merge_cells("B15:C15")
    c = ws["B15"]
    c.value = total_dev
    c.font = Font(bold=True, color=GREEN_NEON, size=28, name="Calibri")
    c.fill = fill(DARK)
    c.alignment = align("center", "center")

    ws.merge_cells("B16:C16")
    c = ws["B16"]
    c.value = "HORAS DESARROLLO"
    c.font = Font(bold=True, color=GREEN_LIGHT, size=9, name="Calibri")
    c.fill = fill(DARK)
    c.alignment = align("center", "center")

    ws.merge_cells("B17:C17")
    ws["B17"].fill = fill(DARK)

    # Stat box 2 - Horas Total
    ws.merge_cells("D15:E15")
    c = ws["D15"]
    c.value = total_final
    c.font = Font(bold=True, color=GREEN_NEON, size=28, name="Calibri")
    c.fill = fill(GREEN_DARK)
    c.alignment = align("center", "center")

    ws.merge_cells("D16:E16")
    c = ws["D16"]
    c.value = "HORAS TOTAL PROYECTO"
    c.font = Font(bold=True, color=GREEN_LIGHT, size=9, name="Calibri")
    c.fill = fill(GREEN_DARK)
    c.alignment = align("center", "center")

    ws.merge_cells("D17:E17")
    ws["D17"].fill = fill(GREEN_DARK)

    # ═══════════════════════════════════════════
    # ACTIVITIES TABLE  (rows 19+)
    # ═══════════════════════════════════════════
    row_h(19, 8)

    # Table title
    tbl_start = 20
    row_h(tbl_start, 28)
    ws.merge_cells(f"B{tbl_start}:E{tbl_start}")
    c = ws.cell(tbl_start, 2, value="  DESGLOSE DE ACTIVIDADES")
    c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    c.fill = fill(GREEN_MED)
    c.alignment = align("left", "center")

    # Column headers
    header_row = tbl_start + 1
    row_h(header_row, 26)
    headers = ["ACTIVIDAD", "DESCRIPCIÓN", "FUNCIONALIDADES", "HORAS"]
    for ci, h in enumerate(headers, 2):
        c = ws.cell(header_row, ci, value=h)
        c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        c.fill = fill(GREEN_DARK)
        c.alignment = align("center", "center", wrap=True)
        c.border = border_thin("all")

    # Activity rows
    data_start = header_row + 1
    for i, act in enumerate(actividades):
        r = data_start + i
        row_h(r, 48)
        bg = GREEN_PALE if i % 2 == 0 else WHITE
        # Actividad
        c = ws.cell(r, 2, value=act.get("actividad", ""))
        c.font = Font(bold=True, color=GREEN_DARK, size=10, name="Calibri")
        c.fill = fill(bg)
        c.alignment = align("left", "center", wrap=True)
        c.border = border_thin("all")
        # Descripcion
        c = ws.cell(r, 3, value=act.get("descripcion", ""))
        c.font = Font(color="444444", size=9, name="Calibri")
        c.fill = fill(bg)
        c.alignment = align("left", "center", wrap=True)
        c.border = border_thin("all")
        # Funcionalidades
        c = ws.cell(r, 4, value=act.get("funcionalidades", ""))
        c.font = Font(color="444444", size=9, name="Calibri")
        c.fill = fill(bg)
        c.alignment = align("left", "center", wrap=True)
        c.border = border_thin("all")
        # Horas
        c = ws.cell(r, 5, value=act.get("horas", 0))
        c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
        c.fill = fill(GREEN_DARK)
        c.alignment = align("center", "center")
        c.border = border_thin("all")

    # Total row
    total_row = data_start + len(actividades)
    row_h(total_row, 30)
    ws.merge_cells(f"B{total_row}:D{total_row}")
    c = ws.cell(total_row, 2, value="TOTAL DESARROLLO")
    c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    c.fill = fill(DARK)
    c.alignment = align("right", "center")
    c.border = border_thin("all")

    c = ws.cell(total_row, 5, value=total_dev)
    c.font = Font(bold=True, color=GREEN_NEON, size=16, name="Calibri")
    c.fill = fill(DARK)
    c.alignment = align("center", "center")
    c.border = border_thin("all")

    # ═══════════════════════════════════════════
    # PERCENTAGES + NOTES  (below table)
    # ═══════════════════════════════════════════
    pct_start = total_row + 2

    row_h(pct_start, 26)
    ws.merge_cells(f"B{pct_start}:E{pct_start}")
    c = ws.cell(pct_start, 2, value="  FACTORES ADICIONALES")
    c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    c.fill = fill(GREEN_MED)
    c.alignment = align("left", "center")

    factors = [
        ("Pruebas Unitarias",    data.get("pruebas_pct", 0)),
        ("Entendimiento",        data.get("entendimiento_pct", 0)),
        ("Riesgo",               data.get("riesgo_pct", 0)),
    ]
    for j, (fname, fval) in enumerate(factors):
        r = pct_start + 1 + j
        row_h(r, 22)
        bg = GREEN_PALE if j % 2 == 0 else WHITE
        ws.merge_cells(f"B{r}:D{r}")
        c = ws.cell(r, 2, value=f"  {fname}")
        c.font = Font(color=DARK, size=10, name="Calibri")
        c.fill = fill(bg)
        c.alignment = align("left", "center")
        c.border = border_thin("all")
        c = ws.cell(r, 5, value=f"{fval}%")
        c.font = Font(bold=True, color=GREEN_DARK, size=10, name="Calibri")
        c.fill = fill(bg)
        c.alignment = align("center", "center")
        c.border = border_thin("all")

    # Grand total
    gt_row = pct_start + 4
    row_h(gt_row, 34)
    ws.merge_cells(f"B{gt_row}:D{gt_row}")
    c = ws.cell(gt_row, 2, value="  TOTAL FINAL DEL PROYECTO")
    c.font = Font(bold=True, color=WHITE, size=12, name="Calibri")
    c.fill = fill(DARK)
    c.alignment = align("left", "center")
    c.border = border_thin("all")

    c = ws.cell(gt_row, 5, value=total_final)
    c.font = Font(bold=True, color=GREEN_NEON, size=16, name="Calibri")
    c.fill = fill(DARK)
    c.alignment = align("center", "center")
    c.border = border_thin("all")

    # Notes
    notas = data.get("notas", [])
    if notas:
        notes_start = gt_row + 2
        row_h(notes_start, 26)
        ws.merge_cells(f"B{notes_start}:E{notes_start}")
        c = ws.cell(notes_start, 2, value="  NOTAS IMPORTANTES")
        c.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        c.fill = fill(GREEN_MED)
        c.alignment = align("left", "center")
        for k, nota in enumerate(notas):
            r = notes_start + 1 + k
            row_h(r, 28)
            ws.merge_cells(f"B{r}:E{r}")
            c = ws.cell(r, 2, value=f"  • {nota}")
            c.font = Font(color="444444", size=9, italic=True, name="Calibri")
            c.fill = fill(GREEN_PALE)
            c.alignment = align("left", "center", wrap=True)
            c.border = Border(bottom=Side(style="thin", color=GRAY_MID))

    # ── Footer row ────────────────────────────────────────────────────
    footer_row = ws.max_row + 2
    row_h(footer_row, 20)
    ws.merge_cells(f"B{footer_row}:E{footer_row}")
    c = ws.cell(footer_row, 2,
        value="Periferia IT Group  •  Desarrollado por Juan Aragón  •  📞 314 674 7578")
    c.font = Font(color="888888", size=9, italic=True, name="Calibri")
    c.alignment = align("center", "center")
    c.fill = fill(GREEN_PALE)

    # ── Freeze panes ─────────────────────────────────────────────────
    ws.freeze_panes = "B1"
    ws.sheet_view.showGridLines = False

    # Guardado temporal
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return send_file(tmp.name, as_attachment=True, download_name="Estimacion_Periferia.xlsx")

if __name__ == "__main__":
    # CONFIGURACIÓN PARA RENDER
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)